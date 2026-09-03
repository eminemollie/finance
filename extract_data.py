#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extract_data.py
從「財務管理系統.xlsx」擷取關鍵資料，輸出 data.json 供手機版網頁讀取。
用途：GitHub Actions 在偵測到 xlsx 更新時自動執行本腳本。

執行方式：python3 extract_data.py <xlsx路徑> <輸出json路徑>
"""
import sys
import json
import datetime
import openpyxl


def col_letter(idx):
    letter = ''
    n = idx
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letter = chr(65 + rem) + letter
    return letter


def to_iso_date(v):
    """Excel 的日期值轉成 YYYY-MM-DD 字串（JS Date 可直接解析）"""
    if v is None or v == '':
        return None
    if isinstance(v, (datetime.date, datetime.datetime)):
        return v.strftime('%Y-%m-%d')
    return None


def get_jiayan_remaining(wb):
    """直接從股票投資分頁的買入/賣出原始記錄重新計算剩餘金額，不依賴任何公式快取"""
    try:
        ws = wb['股票投資']
    except KeyError:
        return None
    buy_total, sell_total = 0, 0
    mode = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        c = ws.cell(row=r, column=3).value
        if a == '買入記錄':
            mode = 'buy'; continue
        if a == '賣出記錄':
            mode = 'sell'; continue
        if isinstance(a, str) and ('小結' in a or '共投入' in a or '共收入' in a):
            mode = None; continue
        if mode == 'buy' and isinstance(a, int) and isinstance(c, (int, float)):
            buy_total += c
        if mode == 'sell' and isinstance(a, int) and isinstance(c, (int, float)):
            sell_total += c
    if buy_total == 0 and sell_total == 0:
        return None
    return round(buy_total - sell_total)


def get_credit_card_total(wb):
    """直接加總信用卡年支出分頁「信用卡明細」區塊的原始消費金額，
    完全不依賴任何公式快取。掃描範圍僅限信用卡明細（固定+非固定）兩個表格，
    遇到「每年龐大固定支出」區塊即停止，避免誤算年度稅費/保費。"""
    try:
        ws = wb['信用卡年支出']
    except KeyError:
        return None
    total = 0
    in_cc_section = False
    for r in range(1, ws.max_row + 1):
        a_val = ws.cell(row=r, column=1).value
        if isinstance(a_val, str) and '信用卡明細' in a_val:
            in_cc_section = True
            continue
        if isinstance(a_val, str) and ('每年龐大固定支出' in a_val or '龐大固定支出' in a_val):
            break  # 信用卡區塊結束，停止掃描
        if not in_cc_section:
            continue
        for c in range(3, 8):  # C~G 欄：玉山/聯邦/中信/國泰/台新
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)):
                total += v
    return abs(round(total)) if total else None


def get_jiayan_remaining(wb):
    """從股票投資分頁的買賣原始明細重新加總「市場上剩餘金額」，
    不依賴任何公式快取（原本是跨分頁公式，快取遺失時會讀不到）。"""
    try:
        ws = wb['股票投資']
    except KeyError:
        return None
    buy_total, sell_total = 0, 0
    mode = None
    for r in range(1, ws.max_row + 1):
        a_val = ws.cell(row=r, column=1).value
        if isinstance(a_val, str) and '買入記錄' in a_val:
            mode = 'buy'; continue
        if isinstance(a_val, str) and '賣出記錄' in a_val:
            mode = 'sell'; continue
        if isinstance(a_val, str) and '共投入金額' in a_val:
            mode = None; continue
        if isinstance(a_val, str) and '共收入金額' in a_val:
            mode = None; continue
        if isinstance(a_val, str) and '小結' in a_val:
            break
        if mode in ('buy', 'sell'):
            c_val = ws.cell(row=r, column=3).value  # 台幣收入／台幣轉匯
            if isinstance(c_val, (int, float)):
                if mode == 'buy':
                    buy_total += c_val
                else:
                    sell_total += c_val
    return round(buy_total - sell_total)


def extract(xlsx_path):
    # data_only=True 讀取快取值（用於數字），但日期/公式結構仍需從貸款表原始儲存格取得
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    result = {
        'generatedAt': datetime.datetime.now().isoformat(),
        'sourceFile': 'financial_management_system.xlsx',
    }

    # ── 收支明細：假設值 + 收入/支出項目 ─────────────────────────
    ws = wb['收支明細']
    # 配息率假設 (D欄第4列附近，尋找有百分比格式的黃底儲存格)
    rate = None
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None and isinstance(c.value, (int, float)) and 0 < c.value < 1:
                if c.fill and c.fill.fgColor and c.fill.fgColor.rgb == 'FFFFFF00':
                    rate = c.value
                    break
        if rate:
            break
    if rate is None:
        rate = 0.11

    income = []
    expense = []
    section = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        d = ws.cell(row=r, column=4).value
        if isinstance(a, str) and '收入' in a and '合計' not in a:
            section = 'income'
            continue
        if isinstance(a, str) and '支出' in a and '合計' not in a:
            section = 'expense'
            continue
        if b in ('項目', None, ''):
            continue
        if b == '信用卡費' and section == 'expense':
            # 信用卡費是跨分頁公式，快取可能為空；一律改用信用卡年支出分頁的原始明細重新加總
            cc_total = get_credit_card_total(wb)
            amt = cc_total if cc_total is not None else (round(abs(c)) if isinstance(c, (int, float)) else 0)
            expense.append({'name': b, 'amt': amt, 'sub': (d or '')[:24]})
            continue
        if not isinstance(c, (int, float)):
            continue
        if section == 'income':
            is_fund = '基金配息' in str(b)
            if not is_fund:
                income.append({'name': b, 'amt': round(c), 'auto': False})
        elif section == 'expense':
            expense.append({'name': b, 'amt': round(abs(c)), 'sub': (d or '')[:24]})
    income.append({'name': '基金配息(估)', 'amt': 0, 'auto': True})

    result['assumptions_rate'] = rate
    result['income'] = income
    result['expense'] = expense

    # ── 貸款總覽：貸款與壽險結構性事實 ─────────────────────────
    ws = wb['貸款總覽']
    loans = []
    insurance = []
    region_map = {'北區': 'north', '東區': 'east', '-': None}
    # 找出貸款區塊（標題列後直到「合計」列）與壽險區塊
    mode = None
    for r in range(1, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        if b == '貸款名稱':
            mode = 'loan_header'
            continue
        if mode == 'loan_header':
            mode = 'loan'
        if b == '合計':
            mode = None
            continue
        if b == '保單名稱':
            mode = 'ins_header'
            continue
        if mode == 'ins_header':
            mode = 'ins'
        if b == '壽險合計':
            mode = None
            continue

        if mode == 'loan' and b:
            region = ws.cell(row=r, column=3).value
            principal = ws.cell(row=r, column=5).value
            rate_v = ws.cell(row=r, column=6).value
            years = ws.cell(row=r, column=7).value
            start_raw = ws.cell(row=r, column=8).value
            if isinstance(principal, (int, float)) and isinstance(rate_v, (int, float)):
                loans.append({
                    'name': b,
                    'region': region_map.get(region, None),
                    'principal': round(principal),
                    'rate': rate_v,
                    'years': years,
                    'start': to_iso_date(start_raw),
                })
        if mode == 'ins' and b:
            region = ws.cell(row=r, column=3).value
            premium = ws.cell(row=r, column=5).value
            years = ws.cell(row=r, column=6).value
            start_raw = ws.cell(row=r, column=8).value
            if isinstance(premium, (int, float)):
                insurance.append({
                    'name': b + '剩餘保費',
                    'region': region_map.get(region, None),
                    'premium': round(premium),
                    'years': years,
                    'start': to_iso_date(start_raw),
                })

    result['loans'] = loans
    result['insurance'] = insurance

    # ── 基金配息紀錄：申購批次 + 月配息歷史 ─────────────────────
    ws = wb['基金配息紀錄']
    batches = []
    mode = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if a == '批次':
            mode = 'batch'
            continue
        if mode == 'batch' and isinstance(a, int):
            desc = ws.cell(row=r, column=3).value
            nav = ws.cell(row=r, column=4).value
            units = ws.cell(row=r, column=5).value
            twd = ws.cell(row=r, column=6).value
            if isinstance(units, (int, float)):
                batches.append({
                    'no': a, 'date': str(b), 'desc': desc,
                    'nav': nav, 'units': units, 'twd': round(twd) if twd else 0,
                })
        if a == '目前部位小結':
            mode = None

    total_units = sum(x['units'] for x in batches)
    total_twd = sum(x['twd'] for x in batches)

    # 建立「累計單位數 -> 累計投入台幣」對照表，供月配息紀錄比對使用
    # 這樣即使 Excel 快取遺失（H欄公式結果讀不到），也能在 Python 端重新推算 cost，不依賴快取
    cum_units, cum_twd = 0.0, 0
    tier_table = []
    for batch in batches:
        cum_units += batch['units']
        cum_twd += batch['twd']
        tier_table.append((round(cum_units, 3), cum_twd))

    def cost_for_units(units_val):
        """依持有單位數比對批次累計表，回傳對應的累計投入成本(TWD)"""
        best = None
        for tier_units, tier_cost in tier_table:
            if abs(units_val - tier_units) < 0.5:
                return tier_cost
            if units_val <= tier_units:
                best = tier_cost
                break
        return best if best is not None else tier_table[-1][1]

    # 月配息歷史（cost 一律由 Python 依單位數重新推算，不讀取 Excel 公式快取）
    hist = []
    mode = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        if a == '年度' and ws.cell(row=r, column=2).value == '月份':
            mode = 'hist'
            continue
        if mode == 'hist':
            if isinstance(c, (int, float)) and b is not None:
                d_val = ws.cell(row=r, column=4).value  # 持有單位數（原始值，非公式）
                f_val = ws.cell(row=r, column=6).value  # 實收台幣（原始值，非公式）
                if isinstance(d_val, (int, float)) and isinstance(f_val, (int, float)):
                    hist.append({
                        'yr': str(a) if a else '',
                        'mo': str(b),
                        'ud': c,
                        'twd': round(f_val),
                        'units': d_val,
                        'cost': cost_for_units(d_val),
                    })
            if a == '統計摘要':
                mode = None

    result['batches'] = batches
    result['totalUnits'] = round(total_units, 3)
    result['totalInvestedTwd'] = total_twd
    result['hist'] = hist

    # ── 資產負債表：市值假設與資產預設值 ─────────────────────
    ws = wb['資產負債表']
    nav_val, fx_val = 74.91, 31.5
    assets = {'north_re': 0, 'east_re': 0, 'stocks': 0, 'jiayan': 0, 'cash': 0}
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        label = b or a  # 部分列（如家妍投資備忘列）標籤因合併儲存格落在A欄
        if label and '最新淨值' in str(label) and isinstance(c, (int, float)):
            nav_val = c
        if label and '目前匯率' in str(label) and isinstance(c, (int, float)):
            fx_val = c
        if b == '北區不動產估值' and isinstance(c, (int, float)):
            assets['north_re'] = round(c)
        if b == '東區不動產估值' and isinstance(c, (int, float)):
            assets['east_re'] = round(c)
        if b and '股票市值' in str(b) and isinstance(c, (int, float)):
            assets['stocks'] = round(c)
        if label and '家妍投資剩餘部位' in str(label):
            jy = get_jiayan_remaining(wb)
            assets['jiayan'] = jy if jy is not None else (round(c) if isinstance(c, (int, float)) else 0)
        if b == '現金／存款' and isinstance(c, (int, float)):
            assets['cash'] = round(c)

    result['nav'] = nav_val
    result['fx'] = fx_val
    result['assetsDefault'] = assets

    return result


if __name__ == '__main__':
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else 'finance_system.xlsx'
    out_path = sys.argv[2] if len(sys.argv) > 2 else 'data.json'
    data = extract(xlsx_path)
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f'已輸出 {out_path}')
    print(f"收入項目: {len(data['income'])}, 支出項目: {len(data['expense'])}")
    print(f"貸款: {len(data['loans'])}, 壽險: {len(data['insurance'])}")
    print(f"申購批次: {len(data['batches'])}, 配息紀錄: {len(data['hist'])}")
